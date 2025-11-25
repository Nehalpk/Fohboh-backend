from fastapi import HTTPException, UploadFile, File
from fastapi.responses import FileResponse
import boto3
import uuid
import os
import io
import pandas as pd
import logging
import psycopg2
import mimetypes
import traceback
from io import BytesIO  # ✅ This fixes the error
import json
from typing import List, Optional, Dict, Any
from pathlib import Path
from .csv_validation import validate_csv_columns, get_column_requirements

from psycopg2 import pool

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# AWS Configuration
BUCKET_NAME = "my-audio-demo"
UPLOAD_BASE_DIR = "uploads/restaurants"

# Initialize S3 client
s3_client = boto3.client(
    's3',
    region_name='us-east-1',
    aws_access_key_id='AKIA6G75DKGI46PCJQHL',
    aws_secret_access_key='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'
)

# Valid categories
VALID_CATEGORIES = ["Inventory", "Labor", "Sales", "Menu"]

# Global connection pool - lazy initialization
DB_POOL = None

def initialize_db_pool():
    """Initialize the database connection pool if not already initialized"""
    global DB_POOL
    if DB_POOL is None:
        DB_POOL = pool.SimpleConnectionPool(
            5, 20,
            host="database-1-snapshot-26-july.cboosuomg0xi.us-east-1.rds.amazonaws.com",
            database="postgres",
            user="postgres",
            password="EqF3XKOz13DX3jE6APrW"
        )
    return DB_POOL

def get_db_connection():
    """Get a database connection from the pool"""
    try:
        initialize_db_pool()
        return DB_POOL.getconn()
    except Exception as e:
        logger.error(f"Error getting database connection: {str(e)}")
        raise HTTPException(status_code=500, detail="Database connection failed")

def return_db_connection(conn):
    """Return a database connection to the pool"""
    try:
        if conn and DB_POOL:
            DB_POOL.putconn(conn)
    except Exception as e:
        logger.error(f"Error returning database connection: {str(e)}")


def get_restaurant_folder_path(restaurant_name: str) -> str:
    """Create sanitized folder path for a restaurant"""
    safe_name = restaurant_name.replace(" ", "_").lower()
    return f"{UPLOAD_BASE_DIR}/{safe_name}"


def get_user_restaurant_path(restaurant_name: str, user_id: int, category: str) -> str:
    """Get the S3 path for a user's restaurant category folder"""
    restaurant_folder = get_restaurant_folder_path(restaurant_name)
    return f"{restaurant_folder}/user_{user_id}/{category}"


def get_user_restaurant_path_id(restaurant_id: int, user_id: int, category: str) -> str:
    """Get the S3 path for a user's restaurant category folder"""
    restaurant_folder = f"{UPLOAD_BASE_DIR}/{restaurant_id}"
    return f"{restaurant_folder}/user_{user_id}/{category}"


def get_created_by_by_email(current_user_email: str):
    """Fetches the 'created_by' value(s) associated with the given email."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT created_by 
            FROM managers 
            WHERE email = %s
        """, (current_user_email,))

        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="User not found")

        created_by_id = result[0]
        created_by_list = [created_by_id]

        cur.execute("""
            SELECT role, created_by 
            FROM managers 
            WHERE id = %s
        """, (created_by_id,))

        upper_result = cur.fetchone()
        if upper_result:
            role, upper_created_by = upper_result
            if role == "Regional Manager" and upper_created_by is not None:
                created_by_list.append(upper_created_by)

        return created_by_list

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        cur.close()
        return_db_connection(conn)


def get_manager_ids_created_by_user(current_user_id: int) -> list[int]:
    """Return a list of manager IDs where created_by = current_user_id"""
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id FROM managers
            WHERE created_by = %s
        """, (current_user_id,))

        rows = cur.fetchall()
        return [row[0] for row in rows]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        cur.close()
        return_db_connection(conn)


def get_user_restaurant_path_r3(restaurant_name: str, current_user: dict) -> list[str]:
    """Get the S3 path(s) for a user's restaurant category folder"""
    restaurant_folder = get_restaurant_folder_path(restaurant_name)

    if current_user["role"] == "SUPER_ADMIN":
        user_res_path = [f"{restaurant_folder}"]

    elif current_user["role"] == "Restaurant Owner":
        user_res_path = [f"{restaurant_folder}/user_{current_user['id']}"]
        manager_ids = get_manager_ids_created_by_user(current_user["id"])
        manager_paths = [f"{restaurant_folder}/user_{manager_id}" for manager_id in manager_ids]
        user_res_path.extend(manager_paths)

    elif current_user["role"] == "Regional Manager":
        user_res_path_1 = f"{restaurant_folder}/user_{current_user['id']}"
        created_by_list = get_created_by_by_email(current_user["email"])

        if not created_by_list:
            raise HTTPException(status_code=404, detail="Created By not found")

        user_res_path_list = [user_res_path_1] + [f"{restaurant_folder}/user_{creator_id}" for creator_id in created_by_list]
        user_res_path = user_res_path_list

    elif current_user["role"] == "Restaurant Manager":
        user_res_path_1 = f"{restaurant_folder}/user_{current_user['id']}"
        created_by_list = get_created_by_by_email(current_user["email"])

        if not created_by_list:
            raise HTTPException(status_code=404, detail="Created By not found")

        user_res_path_list = [user_res_path_1] + [f"{restaurant_folder}/user_{creator_id}" for creator_id in created_by_list]
        user_res_path = user_res_path_list

    else:
        raise HTTPException(status_code=403, detail="Unauthorized role")

    return user_res_path


def get_user_restaurant_path_r3_id(restaurant_id: int, current_user: dict) -> list[str]:
    """Get the S3 path(s) for a user's restaurant category folder"""
    restaurant_folder = f"{UPLOAD_BASE_DIR}/{restaurant_id}"

    if current_user["role"] == "SUPER_ADMIN":
        user_res_path = [f"{restaurant_folder}"]

    elif current_user["role"] == "Restaurant Owner":
        user_res_path = [f"{restaurant_folder}/user_{current_user['id']}"]
        manager_ids = get_manager_ids_created_by_user(current_user["id"])
        manager_paths = [f"{restaurant_folder}/user_{manager_id}" for manager_id in manager_ids]
        user_res_path.extend(manager_paths)

    elif current_user["role"] == "Regional Manager":
        user_res_path_1 = f"{restaurant_folder}/user_{current_user['id']}"
        created_by_list = get_created_by_by_email(current_user["email"])

        if not created_by_list:
            raise HTTPException(status_code=404, detail="Created By not found")

        user_res_path_list = [user_res_path_1] + [f"{restaurant_folder}/user_{creator_id}" for creator_id in created_by_list]
        user_res_path = user_res_path_list

    elif current_user["role"] == "Restaurant Manager":
        user_res_path_1 = f"{restaurant_folder}/user_{current_user['id']}"
        created_by_list = get_created_by_by_email(current_user["email"])

        if not created_by_list:
            raise HTTPException(status_code=404, detail="Created By not found")

        user_res_path_list = [user_res_path_1] + [f"{restaurant_folder}/user_{creator_id}" for creator_id in created_by_list]
        user_res_path = user_res_path_list

    else:
        raise HTTPException(status_code=403, detail="Unauthorized role")

    return user_res_path


async def verify_restaurant_access(restaurant_name: str, current_user: dict) -> dict:
    """Verify user has access to the restaurant and it exists"""
    conn = get_db_connection()
    try:
        cur = conn.cursor()

        if current_user["role"] == "SUPER_ADMIN":
            cur.execute("""
                SELECT id, name, location
                FROM restaurants
                WHERE name = %s AND active = true
            """, (restaurant_name,))

        elif current_user["role"] == "Restaurant Owner":
            cur.execute("""
                SELECT r.id, r.name, r.location
                FROM restaurants r
                LEFT JOIN managers m ON r.created_by = m.id
                WHERE r.name = %s AND r.active = true AND r.created_by = %s
            """, (restaurant_name, current_user["id"]))

        else:
            cur.execute("""
                SELECT r.id, r.name, r.location
                FROM restaurants r
                JOIN restaurant_assignments ra ON r.id = ra.restaurant_id
                WHERE r.name = %s AND r.active = true AND ra.manager_id = %s
            """, (restaurant_name, current_user["id"]))

        restaurant = cur.fetchone()
        if not restaurant:
            raise HTTPException(
                status_code=404,
                detail=f"Restaurant '{restaurant_name}' not found or you don't have access to it"
            )

        return {"id": restaurant[0], "name": restaurant[1], "location": restaurant[2]}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verifying restaurant access: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error verifying restaurant access: {str(e)}")
    finally:
        cur.close()
        return_db_connection(conn)


async def verify_restaurant_access_id(restaurant_id: int, current_user: dict) -> dict:
    """Verify user has access to the restaurant and it exists"""
    conn = get_db_connection()
    try:
        cur = conn.cursor()

        if current_user["role"] == "SUPER_ADMIN":
            cur.execute("""
                SELECT id, name, location
                FROM restaurants
                WHERE id = %s AND active = true
            """, (restaurant_id,))

        elif current_user["role"] == "Restaurant Owner":
            cur.execute("""
                SELECT r.id, r.name, r.location
                FROM restaurants r
                LEFT JOIN managers m ON r.created_by = m.id
                WHERE r.id = %s AND r.active = true AND r.created_by = %s
            """, (restaurant_id, current_user["id"]))

        else:
            cur.execute("""
                SELECT r.id, r.name, r.location
                FROM restaurants r
                JOIN restaurant_assignments ra ON r.id = ra.restaurant_id
                WHERE r.id = %s AND r.active = true AND ra.manager_id = %s
            """, (restaurant_id, current_user["id"]))

        restaurant = cur.fetchone()
        if not restaurant:
            raise HTTPException(
                status_code=404,
                detail=f"Restaurant '{restaurant_id}' not found or you don't have access to it"
            )

        return {"id": restaurant[0], "name": restaurant[1], "location": restaurant[2]}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verifying restaurant access: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error verifying restaurant access: {str(e)}")
    finally:
        cur.close()
        return_db_connection(conn)


async def list_csv_files(restaurant_name: str, current_user: dict) -> List[Dict[str, Any]]:
    """List all files for a restaurant in S3 without filtering by category"""
    try:
        logger.info(f"Listing files for restaurant: {restaurant_name}, user role: {current_user.get('role', 'Unknown')}")
        
        restaurant = await verify_restaurant_access(restaurant_name, current_user)
        logger.info(f"Restaurant access verified for: {restaurant_name}")

        if current_user["role"] == "SUPER_ADMIN":
            restaurant_folder = get_restaurant_folder_path(restaurant_name)
            logger.info(f"SUPER_ADMIN listing files from folder: {restaurant_folder}")
            
            try:
                files = []
                response = s3_client.list_objects_v2(
                    Bucket=BUCKET_NAME,
                    Prefix=f"{restaurant_folder}/user_",
                    Delimiter='/'
                )
                
                user_folders = []
                if 'CommonPrefixes' in response:
                    for prefix_info in response['CommonPrefixes']:
                        prefix = prefix_info['Prefix'].rstrip('/')
                        user_folders.append(prefix)
                
                logger.info(f"Found {len(user_folders)} user folders: {user_folders}")
                
                for user_folder in user_folders:
                    response = s3_client.list_objects_v2(
                        Bucket=BUCKET_NAME,
                        Prefix=f"{user_folder}/"
                    )
                    
                    if 'Contents' in response:
                        for obj in response['Contents']:
                            if obj['Key'].endswith('/'):
                                continue

                            filename = os.path.basename(obj['Key'])

                            if filename:
                                files.append({
                                    "filename": filename,
                                    "size": obj['Size'],
                                    "last_modified": obj['LastModified'].strftime("%Y-%m-%d %H:%M:%S"),
                                    "s3_key": obj['Key'],
                                    "category": str(obj['Key']).split("/")[-2],
                                    "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                                })

                logger.info(f"Found {len(files)} total files for SUPER_ADMIN")
                return {
                    "restaurant": restaurant_name,
                    "files": files,
                    "count": len(files)
                }
                
            except Exception as e:
                logger.error(f"Error listing files from S3 (SUPER_ADMIN): {str(e)}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                raise HTTPException(status_code=500, detail=f"Error retrieving file list: {str(e)}")
        
        else:
            logger.info(f"Regular user listing files for role: {current_user['role']}")
            s3_folder_list = get_user_restaurant_path_r3(restaurant_name, current_user)
            logger.info(f"User folders to search: {s3_folder_list}")

            try:
                files = []
                for s3_folder in s3_folder_list:
                    response = s3_client.list_objects_v2(
                        Bucket=BUCKET_NAME,
                        Prefix=f"{s3_folder}/"
                    )

                    if 'Contents' in response:
                        for obj in response['Contents']:
                            if obj['Key'].endswith('/'):
                                continue

                            filename = os.path.basename(obj['Key'])

                            if filename:
                                files.append({
                                    "filename": filename,
                                    "size": obj['Size'],
                                    "last_modified": obj['LastModified'].strftime("%Y-%m-%d %H:%M:%S"),
                                    "s3_key": obj['Key'],
                                    "category": str(obj['Key']).split("/")[-2],
                                    "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                                })

                logger.info(f"Found {len(files)} total files for regular user")
                return {
                    "restaurant": restaurant_name,
                    "files": files,
                    "count": len(files)
                }

            except Exception as e:
                logger.error(f"Error listing files from S3 (regular user): {str(e)}")
                logger.error(f"Traceback: {traceback.format_exc()}")
                raise HTTPException(status_code=500, detail=f"Error retrieving file list: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in list_csv_files: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred while listing files. Please try again.")


async def list_s3_files_id(restaurant_id: int, current_user: dict) -> List[Dict[str, Any]]:
    """List all files for a restaurant in S3 without filtering by category"""
    try:
        restaurant = await verify_restaurant_access_id(restaurant_id, current_user)

        if current_user["role"] == "SUPER_ADMIN":
            restaurant_folder = f"{UPLOAD_BASE_DIR}/{restaurant_id}"
            
            try:
                files = []
                response = s3_client.list_objects_v2(
                    Bucket=BUCKET_NAME,
                    Prefix=f"{restaurant_folder}/user_",
                    Delimiter='/'
                )
                
                user_folders = []
                if 'CommonPrefixes' in response:
                    for prefix_info in response['CommonPrefixes']:
                        prefix = prefix_info['Prefix'].rstrip('/')
                        user_folders.append(prefix)
                
                for user_folder in user_folders:
                    response = s3_client.list_objects_v2(
                        Bucket=BUCKET_NAME,
                        Prefix=f"{user_folder}/"
                    )
                    
                    if 'Contents' in response:
                        for obj in response['Contents']:
                            if obj['Key'].endswith('/'):
                                continue

                            filename = os.path.basename(obj['Key'])

                            if filename:
                                files.append({
                                    "filename": filename,
                                    "size": obj['Size'],
                                    "last_modified": obj['LastModified'].strftime("%Y-%m-%d %H:%M:%S"),
                                    "s3_key": obj['Key'],
                                    "category": str(obj['Key']).split("/")[-2],
                                    "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                                })

                return {
                    "restaurant": restaurant_id,
                    "files": files,
                    "count": len(files)
                }
                
            except Exception as e:
                logger.error(f"Error listing files from S3: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error retrieving file list: {str(e)}")
        
        else:
            s3_folder_list = get_user_restaurant_path_r3_id(restaurant_id, current_user)

            try:
                files = []
                for s3_folder in s3_folder_list:
                    response = s3_client.list_objects_v2(
                        Bucket=BUCKET_NAME,
                        Prefix=f"{s3_folder}/"
                    )

                    if 'Contents' in response:
                        for obj in response['Contents']:
                            if obj['Key'].endswith('/'):
                                continue

                            filename = os.path.basename(obj['Key'])

                            if filename:
                                files.append({
                                    "filename": filename,
                                    "size": obj['Size'],
                                    "last_modified": obj['LastModified'].strftime("%Y-%m-%d %H:%M:%S"),
                                    "s3_key": obj['Key'],
                                    "category": str(obj['Key']).split("/")[-2],
                                    "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                                })

                return {
                    "restaurant": restaurant_id,
                    "files": files,
                    "count": len(files)
                }

            except Exception as e:
                logger.error(f"Error listing files from S3: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error retrieving file list: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in list_csv_files: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error listing files: {str(e)}")


async def download_csv_file(restaurant_name: str, category: str, filename: str, current_user: dict):
    """Download file"""
    try:
        restaurant = await verify_restaurant_access(restaurant_name, current_user)

        if not filename:
            raise HTTPException(status_code=400, detail="Filename is required")
        
        file_type = filename.split('.')[-1].lower()

        if current_user["role"] == "SUPER_ADMIN":
            restaurant_folder = get_restaurant_folder_path(restaurant_name)
            
            try:
                response = s3_client.list_objects_v2(
                    Bucket=BUCKET_NAME,
                    Prefix=f"{restaurant_folder}/user_",
                    Delimiter='/'
                )
                
                file_content = None
                found_key = None
                
                user_folders = []
                if 'CommonPrefixes' in response:
                    for prefix_info in response['CommonPrefixes']:
                        prefix = prefix_info['Prefix'].rstrip('/')
                        user_folders.append(prefix)
                
                for user_folder in user_folders:
                    s3_key = f"{user_folder}/{category}/{filename}"
                    try:
                        response = s3_client.get_object(Bucket=BUCKET_NAME, Key=s3_key)
                        file_content = response['Body'].read()
                        found_key = s3_key
                        break
                    except s3_client.exceptions.NoSuchKey:
                        continue
                    except Exception as e:
                        logger.error(f"Error accessing S3 key {s3_key}: {str(e)}")
                        continue
                
                if file_content is None:
                    raise HTTPException(status_code=404, detail=f"File '{filename}' not found in any user folder")
                
                return file_content, {
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Content-Type': file_type
                }
                
            except Exception as e:
                logger.error(f"Error searching for file in S3: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error searching for file: {str(e)}")
        
        else:
            s3_folder = get_user_restaurant_path(restaurant_name, current_user["id"], category)
            s3_key = f"{s3_folder}/{filename}"

            try:
                response = s3_client.get_object(Bucket=BUCKET_NAME, Key=s3_key)
                file_content = response['Body'].read()

                return file_content, {
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Content-Type': file_type
                }

            except s3_client.exceptions.NoSuchKey:
                raise HTTPException(status_code=404, detail=f"File '{filename}' not found")
            except Exception as e:
                logger.error(f"Error downloading file from S3: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in download_file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

#----------------------------#
async def delete_csv_file(restaurant_name: str, category: str, filename: str, current_user: dict) -> Dict[str, Any]:
    """Delete a CSV file from both S3 and PostgreSQL"""
    try:
        restaurant = await verify_restaurant_access(restaurant_name, current_user)
        restaurant_id = restaurant["id"]

        if not filename:
            raise HTTPException(status_code=400, detail="Filename is required")

        found_key = None
        
        # Search for the file in accessible folders
        if current_user["role"] == "SUPER_ADMIN":
            restaurant_folder = get_restaurant_folder_path(restaurant_name)
            
            response = s3_client.list_objects_v2(
                Bucket=BUCKET_NAME,
                Prefix=f"{restaurant_folder}/user_",
                Delimiter='/'
            )
            
            user_folders = []
            if 'CommonPrefixes' in response:
                for prefix_info in response['CommonPrefixes']:
                    prefix = prefix_info['Prefix'].rstrip('/')
                    user_folders.append(prefix)
            
            for user_folder in user_folders:
                s3_key = f"{user_folder}/{category}/{filename}"
                try:
                    s3_client.head_object(Bucket=BUCKET_NAME, Key=s3_key)
                    found_key = s3_key
                    break
                except s3_client.exceptions.ClientError:
                    continue
        
        else:
            s3_folder_list = get_user_restaurant_path_r3(restaurant_name, current_user)
            
            for s3_folder in s3_folder_list:
                s3_key = f"{s3_folder}/{category}/{filename}"
                try:
                    s3_client.head_object(Bucket=BUCKET_NAME, Key=s3_key)
                    found_key = s3_key
                    break
                except s3_client.exceptions.ClientError:
                    continue

        if not found_key:
            raise HTTPException(status_code=404, detail=f"File '{filename}' not found")

        # Delete from S3
        s3_client.delete_object(Bucket=BUCKET_NAME, Key=found_key)
        logger.info(f"✅ Deleted file from S3: {found_key}")

        # Delete from PostgreSQL
        conn = get_db_connection()
        deleted_count = 0
        
        try:
            cur = conn.cursor()
            
            if category == 'Sales':
                cur.execute("DELETE FROM sales_graphs WHERE restaurant_id = %s AND filename = %s", 
                           (restaurant_id, filename))
                deleted_count = cur.rowcount
                
            elif category == 'Inventory':
                cur.execute("DELETE FROM inventory_graphs WHERE restaurant_id = %s AND filename = %s", 
                           (restaurant_id, filename))
                deleted_count = cur.rowcount
                
            elif category == 'Menu':
                cur.execute("DELETE FROM menu_graphs WHERE restaurant_id = %s AND filename = %s", 
                           (restaurant_id, filename))
                deleted_count = cur.rowcount
                
            elif category == 'Labor':
                cur.execute("DELETE FROM staff_info WHERE restaurant_id = %s AND filename = %s", 
                           (restaurant_id, filename))
                deleted_count = cur.rowcount
                
                cur.execute("DELETE FROM employees_graphs WHERE restaurant_id = %s AND filename = %s", 
                           (restaurant_id, filename))
                deleted_count += cur.rowcount
            
            conn.commit()
            logger.info(f"✅ Deleted {deleted_count} database records")
            
        finally:
            cur.close()
            return_db_connection(conn)

        return {
            "message": f"File '{filename}' deleted successfully",
            "restaurant": restaurant_name,
            "category": category,
            "filename": filename,
            "s3_deleted": True,
            "db_records_deleted": deleted_count
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in delete_csv_file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting CSV file: {str(e)}")
#-----------------------------#
async def delete_all_user_files(restaurant_name: str, current_user: dict) -> Dict[str, Any]:
    """Delete ALL files for the current user across all categories from both S3 and PostgreSQL"""
    try:
        restaurant = await verify_restaurant_access(restaurant_name, current_user)
        restaurant_id = restaurant["id"]

        categories = ["Sales", "Inventory", "Menu", "Labor"]
        total_s3_deleted = 0
        total_db_deleted = 0
        deletion_details = []

        # Get connection for database operations
        conn = get_db_connection()
        
        try:
            cur = conn.cursor()
            
            for category in categories:
                s3_files_deleted = 0
                db_records_deleted = 0
                
                try:
                    # ============= DELETE FROM S3 =============
                    if current_user["role"] == "SUPER_ADMIN":
                        # SUPER_ADMIN: Delete from all user folders
                        restaurant_folder = get_restaurant_folder_path(restaurant_name)
                        
                        response = s3_client.list_objects_v2(
                            Bucket=BUCKET_NAME,
                            Prefix=f"{restaurant_folder}/user_",
                            Delimiter='/'
                        )
                        
                        user_folders = []
                        if 'CommonPrefixes' in response:
                            for prefix_info in response['CommonPrefixes']:
                                prefix = prefix_info['Prefix'].rstrip('/')
                                user_folders.append(prefix)
                        
                        # Delete files from all user folders
                        for user_folder in user_folders:
                            response = s3_client.list_objects_v2(
                                Bucket=BUCKET_NAME,
                                Prefix=f"{user_folder}/{category}/"
                            )
                            
                            if 'Contents' in response:
                                for obj in response['Contents']:
                                    if obj['Key'].endswith('/'):
                                        continue
                                    
                                    try:
                                        s3_client.delete_object(Bucket=BUCKET_NAME, Key=obj['Key'])
                                        s3_files_deleted += 1
                                        logger.info(f"✅ Deleted from S3: {obj['Key']}")
                                    except Exception as e:
                                        logger.error(f"Error deleting {obj['Key']} from S3: {str(e)}")
                    
                    else:
                        # Regular users: Only delete from their accessible folders
                        s3_folder = get_user_restaurant_path(restaurant_name, current_user["id"], category)
                        
                        response = s3_client.list_objects_v2(
                            Bucket=BUCKET_NAME,
                            Prefix=f"{s3_folder}/"
                        )
                        
                        if 'Contents' in response:
                            for obj in response['Contents']:
                                if obj['Key'].endswith('/'):
                                    continue
                                
                                try:
                                    s3_client.delete_object(Bucket=BUCKET_NAME, Key=obj['Key'])
                                    s3_files_deleted += 1
                                    logger.info(f"✅ Deleted from S3: {obj['Key']}")
                                except Exception as e:
                                    logger.error(f"Error deleting {obj['Key']} from S3: {str(e)}")
                    
                    # ============= DELETE FROM POSTGRESQL =============
                    if category == 'Sales':
                        cur.execute("""
                            DELETE FROM sales_graphs 
                            WHERE restaurant_id = %s
                        """, (restaurant_id,))
                        db_records_deleted = cur.rowcount
                        
                    elif category == 'Inventory':
                        cur.execute("""
                            DELETE FROM inventory_graphs 
                            WHERE restaurant_id = %s
                        """, (restaurant_id,))
                        db_records_deleted = cur.rowcount
                        
                    elif category == 'Menu':
                        cur.execute("""
                            DELETE FROM menu_graphs 
                            WHERE restaurant_id = %s
                        """, (restaurant_id,))
                        db_records_deleted = cur.rowcount
                        
                    elif category == 'Labor':
                        cur.execute("""
                            DELETE FROM staff_info 
                            WHERE restaurant_id = %s
                        """, (restaurant_id,))
                        db_records_deleted = cur.rowcount
                        
                        cur.execute("""
                            DELETE FROM employees_graphs 
                            WHERE restaurant_id = %s
                        """, (restaurant_id,))
                        db_records_deleted += cur.rowcount
                    
                    logger.info(f"✅ Deleted {db_records_deleted} records from {category} table")
                    
                    total_s3_deleted += s3_files_deleted
                    total_db_deleted += db_records_deleted
                    
                    deletion_details.append({
                        "category": category,
                        "s3_files_deleted": s3_files_deleted,
                        "db_records_deleted": db_records_deleted
                    })
                    
                except Exception as e:
                    logger.error(f"Error processing category {category}: {str(e)}")
                    logger.error(f"Traceback: {traceback.format_exc()}")
                    deletion_details.append({
                        "category": category,
                        "error": str(e),
                        "s3_files_deleted": s3_files_deleted,
                        "db_records_deleted": db_records_deleted
                    })
            
            conn.commit()
            logger.info(f"✅ Successfully deleted all files and records for user {current_user['id']} in restaurant {restaurant_name}")
            
        except Exception as e:
            conn.rollback()
            logger.error(f"Error during bulk deletion: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise
        finally:
            cur.close()
            return_db_connection(conn)

        return {
            "message": "All user files and database records deleted successfully",
            "restaurant": restaurant_name,
            "user_id": current_user["id"],
            "restaurant_id": restaurant_id,
            "total_s3_files_deleted": total_s3_deleted,
            "total_db_records_deleted": total_db_deleted,
            "details_by_category": deletion_details
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in delete_all_user_files: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error deleting all user files: {str(e)}")
# File validation and processing functions
async def validate_file(file: UploadFile):
    """Validate file type (PDF, XLS, DOCX, or CSV)"""
    file_type = file.filename.split('.')[-1].lower()

    if file_type not in ["pdf", "csv", "xls", "xlsx", "docx"]:
        raise HTTPException(status_code=400, detail="File must be a PDF, CSV, XLSX, or DOCX")

    return file_type


async def process_csv(file: UploadFile):
    """Process CSV file and return row/column count"""
    file_content = await file.read()
    try:
        df = pd.read_csv(io.BytesIO(file_content))
        return len(df), len(df.columns)
    except Exception as e:
        logger.error(f"Invalid CSV format: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Invalid CSV format: {str(e)}")


async def process_xls(file: UploadFile):
    """Process XLSX file and return row/column count"""
    from openpyxl import load_workbook
    file_content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        sheet = wb.active
        row_count = sheet.max_row
        column_count = sheet.max_column
        return row_count, column_count
    except Exception as e:
        logger.error(f"Invalid XLS format: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Invalid XLS format: {str(e)}")


async def process_docx(file: UploadFile):
    """Process DOCX file and return the number of paragraphs"""
    from docx import Document
    file_content = await file.read()
    try:
        doc = Document(io.BytesIO(file_content))
        paragraph_count = len(doc.paragraphs)
        return paragraph_count, 0
    except Exception as e:
        logger.error(f"Invalid DOCX format: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Invalid DOCX format: {str(e)}")


async def process_pdf(file: UploadFile):
    """Process PDF file and return the number of pages"""
    from PyPDF2 import PdfReader
    file_content = await file.read()
    try:
        pdf = PdfReader(io.BytesIO(file_content))
        page_count = len(pdf.pages)
        return page_count, 1
    except Exception as e:
        logger.error(f"Invalid PDF format: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Invalid PDF format: {str(e)}")


CUSTOM_MIME_TYPES = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv": "text/csv",
    ".pdf": "application/pdf"
}


def get_content_type(filename: str) -> str:
    ext = filename.lower().split('.')[-1]
    if ext and f".{ext}" in CUSTOM_MIME_TYPES:
        return CUSTOM_MIME_TYPES[f".{ext}"]

    content_type, _ = mimetypes.guess_type(filename)
    return content_type if content_type else "application/octet-stream"


def get_restaurant_id_by_name(restaurant_name):
    """Fetch the restaurant ID by its name"""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id FROM restaurants
                WHERE name = %s AND active = TRUE
                LIMIT 1
            """, (restaurant_name,))

            result = cur.fetchone()

            if result:
                logging.info("Restaurant '%s' found with ID: %s", restaurant_name, result[0])
                return result[0]
            else:
                logging.warning("Restaurant '%s' not found or is inactive.", restaurant_name)
                return None

    except Exception as e:
        logging.error("Error retrieving restaurant ID for '%s': %s", restaurant_name, str(e))
        logging.debug(traceback.format_exc())
        return None
    finally:
        return_db_connection(conn)


def get_restaurant_name_by_id(restaurant_id):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT name FROM restaurants
                WHERE id = %s AND active = TRUE
            """, (restaurant_id,))

            result = cur.fetchone()

            if result:
                logging.info("Restaurant with ID '%s' found: %s", restaurant_id, result[0])
                return result[0]
            else:
                logging.warning("Restaurant with ID '%s' not found or is inactive.", restaurant_id)
                return None

    except Exception as e:
        logging.error("Error retrieving restaurant ID for '%s': %s", restaurant_id, str(e))
        logging.debug(traceback.format_exc())
        return None
    finally:
        return_db_connection(conn)


# Simplified CSV Processor - No fraud detection
class SimplifiedCSVProcessor:
    """Simplified CSV processor that just processes data without fraud detection"""
    
    async def process_sales_csv(self, file: UploadFile):
        """Process sales CSV file"""
        try:
            await file.seek(0)
            contents = await file.read()
            df = pd.read_csv(io.BytesIO(contents))
    
            # Strip whitespace and normalize column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    
            logging.info(f"Sales CSV columns: {df.columns.tolist()}")
            logging.info(f"Sample row: {df.iloc[0].to_dict() if len(df) > 0 else 'No data'}")
    
            # Helper function to parse dates
            def parse_date(date_str):
                if pd.isna(date_str):
                    return None
                date_str = str(date_str).strip()
                if not date_str or date_str == '':
                    return None
                try:
                    # Try multiple date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            parsed = pd.to_datetime(date_str, format=fmt, errors='coerce')
                            if pd.notna(parsed):
                                return parsed
                        except:
                            continue
                    # Let pandas infer
                    return pd.to_datetime(date_str, errors='coerce')
                except:
                    return None
    
            # Process Date
            if 'date' in df.columns:
                df['date'] = df['date'].apply(parse_date)
                logging.info(f"Parsed Dates - Sample: {df['date'].head()}")
    
            # Process Time
            if 'time' in df.columns:
                df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S', errors='coerce').dt.time
                logging.info(f"Parsed Times - Sample: {df['time'].head()}")
    
            # Process Sale ID
            if 'sale_id' in df.columns:
                df['sale_id'] = df['sale_id'].astype(str)
    
            # Process Items Sold
            if 'items_sold' in df.columns:
                df['items_sold'] = df['items_sold'].astype(str)
    
            # Process numeric columns
            numeric_columns = ['number_of_items', 'subtotal', 'tip', 'total_amount', 
                              'discount_percent', 'covers']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
            # Process Employee ID
            if 'employee_id' in df.columns:
                df['employee_id'] = df['employee_id'].astype(str)
    
            # Process boolean - Is Loyalty Member
            if 'is_loyalty_member' in df.columns:
                df['is_loyalty_member'] = df['is_loyalty_member'].astype(str).str.upper().isin(['TRUE', '1', 'YES'])
    
            # Process string columns
            string_columns = ['promotion_id', 'order_type', 'payment_method']
            for col in string_columns:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str)
    
            logging.info(f"✅ Processed sales CSV successfully: {len(df)} rows")
            return df
            
        except Exception as e:
            logging.error(f"❌ Error processing sales CSV: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            return None

    #@staticmethod
    async def process_inventory_csv(self, file: UploadFile, restaurant_id: int, conn):
        """Process inventory CSV file and insert directly into database"""
        try:
            await file.seek(0)
            contents = await file.read()
            df = pd.read_csv(io.BytesIO(contents))
    
            # Strip whitespace and normalize column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
    
            logging.info(f"📊 Inventory CSV columns: {df.columns.tolist()}")
            logging.info(f"📊 Sample row: {df.iloc[0].to_dict() if len(df) > 0 else 'No data'}")
    
            # Helper function to parse dates
            def parse_date(date_str):
                if pd.isna(date_str):
                    return None
                date_str = str(date_str).strip()
                if not date_str or date_str == '':
                    return None
                try:
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            parsed = pd.to_datetime(date_str, format=fmt, errors='coerce')
                            if pd.notna(parsed):
                                return parsed
                        except:
                            continue
                    return pd.to_datetime(date_str, errors='coerce')
                except:
                    return None
    
            # Process Date
            if 'date' in df.columns:
                df['date'] = df['date'].apply(parse_date)
    
            # Process Last Ordered Date
            if 'last_ordered_date' in df.columns:
                df['last_ordered_date'] = df['last_ordered_date'].apply(parse_date)
    
            # Process Ingredient
            if 'ingredient' in df.columns:
                df['ingredient'] = df['ingredient'].astype(str).str.strip()
    
            # Process numeric columns
            if 'quantity' in df.columns:
                df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0.0)
    
            if 'par_level' in df.columns:
                df['par_level'] = pd.to_numeric(df['par_level'], errors='coerce').fillna(0.0)
    
            if 'unit_cost' in df.columns:
                df['unit_cost'] = pd.to_numeric(df['unit_cost'], errors='coerce').fillna(0.0)
    
            # Process Is Low boolean
            if 'is_low' in df.columns:
                df['is_low'] = df['is_low'].astype(str).str.upper().isin(['TRUE', '1', 'YES'])
                df['alert'] = df['is_low'].apply(lambda x: 'low' if x else 'normal')
            else:
                df['is_low'] = False
                df['alert'] = 'normal'
    
            # Calculate total value
            if 'quantity' in df.columns and 'unit_cost' in df.columns:
                df['total_value'] = df['quantity'] * df['unit_cost']
            else:
                df['total_value'] = 0.0
    
            # Calculate inventory ratio
            if 'quantity' in df.columns and 'par_level' in df.columns:
                df['inventory_ratio'] = df.apply(
                    lambda row: (row['quantity'] / row['par_level']) if row['par_level'] > 0 else 0.0,
                    axis=1
                )
            else:
                df['inventory_ratio'] = 0.0
    
            # Process string columns
            string_columns = ['unit_of_measure', 'supplier']
            for col in string_columns:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str)
    
            logging.info(f"✅ Processed inventory CSV successfully: {len(df)} rows")
            logging.info(f"📊 Low stock items: {df['is_low'].sum()}")
            
            return df
            
        except Exception as e:
            logging.error(f"❌ Error processing inventory CSV: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def process_menu_csv(self, file: UploadFile):
        """Process menu CSV file"""
        try:
            await file.seek(0)
            contents = await file.read()
            df = pd.read_csv(io.BytesIO(contents))
            
            # Strip whitespace and normalize column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            logging.info(f"Menu CSV columns: {df.columns.tolist()}")
            logging.info(f"Sample row: {df.iloc[0].to_dict() if len(df) > 0 else 'No data'}")
            
            # Helper function to parse dates
            def parse_date(date_str):
                if pd.isna(date_str):
                    return None
                date_str = str(date_str).strip()
                if not date_str or date_str == '':
                    return None
                try:
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            parsed = pd.to_datetime(date_str, format=fmt, errors='coerce')
                            if pd.notna(parsed):
                                return parsed
                        except:
                            continue
                    return pd.to_datetime(date_str, errors='coerce')
                except:
                    return None
            
            # Process Created At date
            if 'created_at' in df.columns:
                df['created_at'] = df['created_at'].apply(parse_date)
                df['date'] = df['created_at']  # Also store as 'date' for consistency
                logging.info(f"Parsed Created At - Sample: {df['created_at'].head()}")
            elif 'date' in df.columns:
                df['date'] = df['date'].apply(parse_date)
            else:
                df['date'] = pd.Timestamp.now()
            
            # Process Menu Item
            if 'menu_item' in df.columns:
                df['menu_item'] = df['menu_item'].astype(str).str.strip()
            else:
                df['menu_item'] = 'Unknown Item'
            
            # Process Ingredient
            if 'ingredient' in df.columns:
                df['ingredient'] = df['ingredient'].astype(str).str.strip()
            else:
                df['ingredient'] = 'Unknown Ingredient'
            
            # Process Amount
            if 'amount' in df.columns:
                df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0.0)
            else:
                df['amount'] = 0.0
            
            # Process Unit Cost
            if 'unit_cost' in df.columns:
                df['unit_cost'] = pd.to_numeric(df['unit_cost'], errors='coerce').fillna(0.0)
            else:
                df['unit_cost'] = 0.0
    
            # Calculate total cost for the ingredient
            df['total_cost'] = df['amount'] * df['unit_cost']
            
            # Process Category
            if 'category' in df.columns:
                df['category'] = df['category'].fillna('').astype(str)
            
            # Process Item Price
            if 'item_price' in df.columns:
                df['item_price'] = pd.to_numeric(df['item_price'], errors='coerce').fillna(0.0)
            
            # Process Is Active
            if 'is_active' in df.columns:
                df['is_active'] = df['is_active'].astype(str).str.upper().isin(['TRUE', '1', 'YES'])
            
            logging.info(f"✅ Processed menu CSV successfully: {len(df)} rows")
            logging.info(f"Unique menu items: {df['menu_item'].nunique()}")
            return df
            
        except Exception as e:
            logging.error(f"❌ Error processing menu CSV: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            return None
        
    async def process_employees_csv(self, file: UploadFile):
        """Process employee CSV file"""
        try:
            await file.seek(0)
            contents = await file.read()
            df = pd.read_csv(io.BytesIO(contents))
            
            # Strip whitespace and normalize column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            logging.info(f"Labor CSV columns: {df.columns.tolist()}")
            logging.info(f"Sample row: {df.iloc[0].to_dict() if len(df) > 0 else 'No data'}")
            
            # Helper function to parse dates
            def parse_date(date_str):
                if pd.isna(date_str):
                    return None
                date_str = str(date_str).strip()
                if not date_str or date_str == '':
                    return None
                try:
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            parsed = pd.to_datetime(date_str, format=fmt, errors='coerce')
                            if pd.notna(parsed):
                                return parsed
                        except:
                            continue
                    return pd.to_datetime(date_str, errors='coerce')
                except:
                    return None
    
            # Helper function to parse time
            def parse_time(time_str):
                if pd.isna(time_str):
                    return None
                try:
                    return pd.to_datetime(time_str, format='%H:%M:%S', errors='coerce').time()
                except:
                    return None
    
            # Process Date
            if 'date' in df.columns:
                df['date'] = df['date'].apply(parse_date)
                logging.info(f"Parsed Dates - Sample: {df['date'].head()}")
    
            # Process Employee ID
            if 'employee_id' in df.columns:
                df['employee_id'] = df['employee_id'].astype(str)
                logging.info(f"Employee IDs - Sample: {df['employee_id'].head()}")
    
            # Process Name
            if 'name' in df.columns:
                df['name'] = df['name'].astype(str).str.strip()
                logging.info(f"Names - Sample: {df['name'].head()}")
            else:
                df['name'] = 'Unknown'
    
            # Process Position (also store as Role)
            if 'position' in df.columns:
                df['position'] = df['position'].astype(str).str.strip()
                df['role'] = df['position']  # Store as role too
                logging.info(f"Positions - Sample: {df['position'].head()}")
            elif 'role' in df.columns:
                df['role'] = df['role'].astype(str).str.strip()
                df['position'] = df['role']
            else:
                df['position'] = 'Unknown'
                df['role'] = 'Unknown'
    
            # Process Hourly Rate
            if 'hourly_rate' in df.columns:
                df['hourly_rate'] = pd.to_numeric(df['hourly_rate'], errors='coerce').fillna(0.0)
                logging.info(f"Hourly Rates - Sample: {df['hourly_rate'].head()}")
            else:
                df['hourly_rate'] = 0.0
    
            # Process Hours Worked
            if 'hours_worked' in df.columns:
                df['hours_worked'] = pd.to_numeric(df['hours_worked'], errors='coerce').fillna(0.0)
                logging.info(f"Hours Worked - Sample: {df['hours_worked'].head()}")
                logging.info(f"Hours Worked range: min={df['hours_worked'].min()}, max={df['hours_worked'].max()}")
            else:
                df['hours_worked'] = 0.0
    
            # Calculate Total Wages
            df['total_wages'] = df['hourly_rate'] * df['hours_worked']
            logging.info(f"Calculated Total Wages - Sample: {df['total_wages'].head()}")
    
            # Process Is Overtime
            if 'is_overtime' in df.columns:
                df['is_overtime'] = df['is_overtime'].astype(str).str.upper().isin(['TRUE', '1', 'YES'])
            else:
                df['is_overtime'] = df['hours_worked'] > 8
    
            # Calculate Overtime Hours
            df['overtime_hours'] = df.apply(
                lambda row: max(0, row['hours_worked'] - 8) if row['is_overtime'] else 0,
                axis=1
            )
    
            # Process Is Salaried
            if 'is_salaried' in df.columns:
                df['is_salaried'] = df['is_salaried'].astype(str).str.upper().isin(['TRUE', '1', 'YES'])
    
            # Process Shift Start
            if 'shift_start' in df.columns:
                df['shift_start'] = df['shift_start'].apply(parse_time)
    
            # Process Shift End
            if 'shift_end' in df.columns:
                df['shift_end'] = df['shift_end'].apply(parse_time)
    
            # Process Department
            if 'department' in df.columns:
                df['department'] = df['department'].fillna('').astype(str)
    
            logging.info(f"✅ Processed labor CSV successfully: {len(df)} rows")
            logging.info(f"📊 Summary:")
            logging.info(f"  - Total Hours: {df['hours_worked'].sum():.2f}")
            logging.info(f"  - Total Wages: ${df['total_wages'].sum():.2f}")
            logging.info(f"  - Overtime Hours: {df['overtime_hours'].sum():.2f}")
            
            return df
            
        except Exception as e:
            logging.error(f"❌ Error processing labor CSV: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            return pd.DataFrame()

    #async def process_employees_csv(self, file: UploadFile):
    #    """Process employee CSV file"""
    #    try:
    #        await file.seek(0)
    #        contents = await file.read()
    #        df = pd.read_csv(io.BytesIO(contents))
    #        logging.info(f"EXACT CSV columns: {df.columns.tolist()}")
    #        logging.info(f"Column types: {df.dtypes.to_dict()}")
    #        for col in ['comps', 'voids', 'tax', 'promos']:
    #            if col in df.columns:
    #                logging.info(f"✓ Found column '{col}' with sample value: {df[col].iloc[0] if len(df) > 0 else 'N/A'}")
    #            else:
    #                logging.warning(f"✗ Column '{col}' NOT FOUND in CSV")
    #
    #        required_columns = ['Employee ID', 'Name']
    #        missing_required = [col for col in required_columns if col not in df.columns]
    #        if missing_required:
    #            logging.warning(f"Missing required columns in employees CSV: {missing_required}")
    #            return pd.DataFrame()
    #
    #        # Helper function to parse dates in DD/MM/YYYY format
    #        def parse_date(date_str):
    #            if pd.isna(date_str):
    #                return None
    #            
    #            # Convert to string and clean it
    #            date_str = str(date_str).strip()
    #            
    #            # If empty, return None
    #            if not date_str or date_str == '':
    #                return None
    #            
    #            # Handle cases where there might be multiple dates (take the first one)
    #            # Example: "23/05/2025 01/15/25" -> take "23/05/2025"
    #            if ' ' in date_str:
    #                date_str = date_str.split()[0]
    #            
    #            try:
    #                # Try parsing DD/MM/YYYY format
    #                parsed_date = pd.to_datetime(date_str, format='%d/%m/%Y', errors='coerce')
    #                if pd.notna(parsed_date):
    #                    return parsed_date
    #                
    #                # If that fails, try automatic parsing
    #                parsed_date = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
    #                return parsed_date
    #            except:
    #                logging.warning(f"Could not parse date: {date_str}")
    #                return None
    #
            # Process Hire Date
            if 'Hire Date' in df.columns:
                df['Hire Date'] = df['Hire Date'].apply(parse_date)
                logging.info(f"Parsed Hire Dates - Sample: {df['Hire Date'].head()}")
            else:
                df['Hire Date'] = pd.NaT
    
            # Process Termination Date
            if 'Termination Date' in df.columns:
                df['Termination Date'] = df['Termination Date'].apply(parse_date)
                logging.info(f"Parsed Termination Dates - Sample: {df['Termination Date'].head()}")
            else:
                df['Termination Date'] = pd.NaT
    
            # Process Hourly Rate
            if 'Hourly Rate' in df.columns:
                df['Hourly Rate'] = pd.to_numeric(df['Hourly Rate'], errors='coerce')
            else:
                df['Hourly Rate'] = 0.0
    
            # Process Role
            if 'Role' not in df.columns:
                df['Role'] = 'Unknown'
    
            # Keep all available columns
            available_columns = [col for col in ['Employee ID', 'Name', 'Role', 'Hire Date', 'Termination Date', 'Hourly Rate']
                                 if col in df.columns]
            processed_df = df[available_columns]
    
            # Log some debug info
            logging.info(f"Processed employees CSV successfully: {len(processed_df)} rows")
            logging.info(f"Date columns summary:")
            logging.info(f"  Hire Date non-null count: {processed_df['Hire Date'].notna().sum()}")
            logging.info(f"  Termination Date non-null count: {processed_df['Termination Date'].notna().sum()}")
            
            return processed_df
            
        except Exception as e:
            logging.error(f"Error processing employees CSV: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            return pd.DataFrame()


# Database uploader
class CSVUploader:
    """Upload processed CSV data to database"""
    
    def __init__(self, conn):
        self.conn = conn
    

    @staticmethod
    def convert_to_json_serializable(value):
        """Convert pandas/numpy types to JSON-serializable types"""
        from datetime import datetime, date, time
        import numpy as np
        
        if pd.isna(value):
            return None
        elif isinstance(value, (pd.Timestamp, datetime, date, time)):
            return value.isoformat()
        elif isinstance(value, (np.integer, np.floating)):
            return value.item()
        elif isinstance(value, np.bool_):
            return bool(value)
        elif isinstance(value, np.ndarray):
            return value.tolist()
        elif hasattr(value, 'item'):  # other numpy types
            return value.item()
        else:
            return value

    def insert_sales(self, processed_df, restaurant_id, filename):
        try:
            import json
            from datetime import datetime, date, time
            
            with self.conn.cursor() as cur:
                query = """
                    INSERT INTO sales_graphs (
                        restaurant_id,
                        filename,
                        data,
                        date,
                        sale_id,
                        time,
                        items_sold,
                        number_of_items,
                        subtotal,
                        tip,
                        total_amount,
                        payment_method,
                        order_type,
                        employee_id,
                        customer_id,
                        is_loyalty_member,
                        promotion_id,
                        discount_applied,
                        table_number,
                        guest_count,
                        seated_time,
                        departure_time,
                        was_reservation,
                        order_received_time,
                        kitchen_start_time,
                        had_modifications,
                        order_accuracy,
                        daypart,
                        comps,
                        voids,
                        tax,
                        promos,
                        seat_time,
                        close_time,
                        service_type,
                        error_flag,
                        fired_time,
                        completed_time,
                        prior_year_net_sales,
                        promo_cost
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """
                
                records = []
                for _, row in processed_df.iterrows():
                    row_data = {key: self.convert_to_json_serializable(value) 
                               for key, value in row.to_dict().items()}
                    
                    # Helper function to safely get values
                    def get_val(col_name, default=None):
                        return row.get(col_name, default) if col_name in row and pd.notna(row[col_name]) else default
                    
                    # Helper for boolean conversion
                    def get_bool(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        return bool(val) if not isinstance(val, str) else val.lower() in ['true', '1', 'yes']
                    
                    # IMPROVED: Helper for datetime conversion
                    def get_datetime(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val
                        # Parse string datetime
                        if isinstance(val, str):
                            try:
                                return pd.to_datetime(val, errors='coerce')
                            except:
                                return None
                        return None
                    
                    # IMPROVED: Helper for time conversion
                    def get_time(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        if isinstance(val, time):
                            return val
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val.time()
                        # Parse string time
                        if isinstance(val, str):
                            try:
                                parsed = pd.to_datetime(val, format='%H:%M:%S', errors='coerce')
                                if pd.notna(parsed):
                                    return parsed.time()
                            except:
                                pass
                        return None
                    
                    # Helper for date conversion
                    def get_date(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None
                    
                    # Extract discount_percent
                    discount_percent = float(get_val('discount_percent', 0.0))
                    subtotal = float(get_val('subtotal', 0.0))
                    discount_applied = (subtotal * discount_percent / 100) if discount_percent > 0 else 0.0
                    
                    records.append((
                        restaurant_id,                                      # 1
                        filename,                                           # 2
                        json.dumps(row_data),                              # 3 - data (jsonb)
                        get_date('date'),                                  # 4 - date
                        str(get_val('sale_id', '')),                       # 5 - sale_id
                        get_time('time'),                                  # 6 - time
                        str(get_val('items_sold', '')),                    # 7 - items_sold
                        int(get_val('number_of_items', 0)) if get_val('number_of_items') is not None else 0,  # 8
                        float(get_val('subtotal', 0.0)),                   # 9 - subtotal
                        float(get_val('tip', 0.0)),                        # 10 - tip
                        float(get_val('total_amount', 0.0)),               # 11 - total_amount
                        str(get_val('payment_method', '')),                # 12 - payment_method
                        str(get_val('order_type', '')),                    # 13 - order_type
                        int(get_val('employee_id')) if get_val('employee_id') is not None else None,      # 14
                        int(get_val('customer_id')) if get_val('customer_id') is not None else None,      # 15
                        get_bool('is_loyalty_member'),                     # 16 - is_loyalty_member
                        int(get_val('promotion_id')) if get_val('promotion_id') is not None else None,    # 17
                        discount_applied,                                  # 18 - discount_applied (calculated)
                        str(get_val('table_number', '')),                  # 19 - table_number
                        int(get_val('covers', 0)) if get_val('covers') is not None else None,  # 20 - guest_count
                        get_datetime('seated_time'),                       # 21 - seated_time (timestamp) ✅
                        get_datetime('departure_time'),                    # 22 - departure_time (timestamp) ✅
                        get_bool('was_reservation'),                       # 23 - was_reservation
                        get_datetime('order_received_time'),               # 24 - order_received_time ✅
                        get_datetime('kitchen_start_time'),                # 25 - kitchen_start_time (timestamp) ✅
                        get_bool('had_modifications'),                     # 26 - had_modifications
                        str(get_val('order_accuracy', '')),                # 27 - order_accuracy
                        str(get_val('daypart', '')),                       # 28 - daypart
                        float(get_val('comps', 0.0)),                      # 29 - comps
                        float(get_val('voids', 0.0)),                      # 30 - voids
                        float(get_val('tax', 0.0)),                        # 31 - tax
                        float(get_val('promos', 0.0)),                     # 32 - promos
                        get_time('seat_time'),                             # 33 - seat_time (time) ✅
                        get_time('close_time'),                            # 34 - close_time (time) ✅
                        str(get_val('service_type', '')),                  # 35 - service_type
                        get_bool('error_flag'),                            # 36 - error_flag
                        get_time('fired_time'),                            # 37 - fired_time (time) ✅
                        get_time('completed_time'),                        # 38 - completed_time (time) ✅
                        float(get_val('prior_year_net_sales', 0.0)),       # 39 - prior_year_net_sales
                        float(get_val('promo_cost', 0.0))                  # 40 - promo_cost
                    ))
                
                cur.executemany(query, records)
                self.conn.commit()
                logging.info(f"✅ Successfully inserted {len(records)} sales records into the database.")
                
                if records:
                    logging.info(f"📊 Sample inserted - Sale ID: {records[0][4]}, Date: {records[0][3]}, Total: ${records[0][10]}")
                
        except Exception as e:
            self.conn.rollback()
            logging.error(f"❌ Error inserting sales data: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
#    
#total_value (quantity × unit_cost)
#inventory_ratio (quantity ÷ par_level)

    def insert_inventory(self, processed_df, restaurant_id, filename):
        try:
            import json
            from datetime import datetime, date
            
            with self.conn.cursor() as cur:
                query = """
                    INSERT INTO inventory_graphs (
                        restaurant_id,
                        filename,
                        data,
                        date,
                        ingredient,
                        quantity,
                        unit,
                        par_level,
                        unit_cost,
                        total_value,
                        category,
                        inventory_ratio,
                        alert,
                        is_low,
                        supplier,
                        last_ordered_date
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """
                
                records = []
                for _, row in processed_df.iterrows():
                    # Convert all values to JSON-serializable for the data column
                    row_data = {key: self.convert_to_json_serializable(value) 
                               for key, value in row.to_dict().items()}
                    
                    # Helper function to safely get values
                    def get_val(col_name, default=None):
                        return row.get(col_name, default) if col_name in row and pd.notna(row[col_name]) else default
                    
                    # Helper for date conversion
                    def get_date(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None
                    
                    records.append((
                        restaurant_id,                              # 1 - restaurant_id
                        filename,                                   # 2 - filename
                        json.dumps(row_data),                      # 3 - data (jsonb)
                        get_date('date'),                          # 4 - date
                        str(get_val('ingredient', '')),            # 5 - ingredient
                        float(get_val('quantity', 0.0)),           # 6 - quantity
                        str(get_val('unit_of_measure', '')),       # 7 - unit
                        float(get_val('par_level', 0.0)),          # 8 - par_level
                        float(get_val('unit_cost', 0.0)),          # 9 - unit_cost
                        float(get_val('total_value', 0.0)),        # 10 - total_value
                        str(get_val('category', '')),              # 11 - category
                        float(get_val('inventory_ratio', 0.0)),    # 12 - inventory_ratio
                        str(get_val('alert', 'normal')),           # 13 - alert
                        bool(get_val('is_low', False)),            # 14 - is_low
                        str(get_val('supplier', '')),              # 15 - supplier
                        get_date('last_ordered_date')              # 16 - last_ordered_date
                    ))
                
                cur.executemany(query, records)
                self.conn.commit()
                logging.info(f"✅ Successfully inserted {len(records)} inventory records into the database.")
                logging.info(f"📊 Low stock items: {sum(1 for r in records if r[14])}")
            # 
            #
        except Exception as e:
            self.conn.rollback()
            logging.error(f"❌ Error inserting inventory data: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")


    def insert_menu(self, processed_df, restaurant_id, filename):
        try:
            import json
            from datetime import datetime, date
            
            with self.conn.cursor() as cur:
                query = """
                    INSERT INTO menu_graphs (
                        restaurant_id,
                        filename,
                        data,
                        date,
                        item_name,
                        category,
                        description,
                        unit_cost,
                        selling_price,
                        profit_margin,
                        recipe,
                        ingredients,
                        amount
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """
                
                records = []
                for _, row in processed_df.iterrows():
                    row_data = {key: self.convert_to_json_serializable(value) 
                               for key, value in row.to_dict().items()}
                    
                    def get_val(col_name, default=None):
                        return row.get(col_name, default) if col_name in row and pd.notna(row[col_name]) else default
                    
                    def get_date(col_name):
                        val = get_val(col_name)
                        if val is None or pd.isna(val):
                            return None
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None
                    
                    # Get values
                    menu_item = str(get_val('menu_item', ''))
                    ingredient = str(get_val('ingredient', ''))
                    unit_cost = float(get_val('unit_cost', 0.0))
                    amount = float(get_val('amount', 0.0))
                    item_price = float(get_val('item_price', 0.0))
                    
                    # Calculate total cost and profit margin
                    total_cost = unit_cost * amount
                    profit_margin = ((item_price - total_cost) / item_price * 100) if item_price > 0 else 0.0
                    
                    records.append((
                        restaurant_id,                           # 1
                        filename,                                # 2
                        json.dumps(row_data),                   # 3 - data
                        get_date('created_at') or get_date('date'),  # 4 - date
                        menu_item,                               # 5 - item_name
                        str(get_val('category', '')),           # 6 - category
                        str(get_val('description', '')),        # 7 - description
                        unit_cost,                               # 8 - unit_cost
                        item_price,                              # 9 - selling_price
                        profit_margin,                           # 10 - profit_margin
                        str(get_val('recipe', '')),             # 11 - recipe
                        ingredient,                              # 12 - ingredients (single ingredient)
                        amount                                   # 13 - amount
                    ))
                
                cur.executemany(query, records)
                self.conn.commit()
                logging.info(f"✅ Successfully inserted {len(records)} menu records into the database.")
                logging.info(f"📊 Unique items: {len(set(r[4] for r in records))}")
                
        except Exception as e:
            self.conn.rollback()
            logging.error(f"❌ Error inserting menu data: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
    

    def insert_employees(self, processed_df, restaurant_id, filename):
        try:
            import json
            from datetime import datetime, date
            
            with self.conn.cursor() as cur:
                query = """
                    INSERT INTO employees_graphs (
                        restaurant_id,
                        filename,
                        data,
                        date,
                        employee_id,
                        employee_name,
                        name,
                        position,
                        role,
                        hourly_rate,
                        hours_worked,
                        total_wages,
                        overtime_hours,
                        hire_date,
                        termination_date
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                """
                
                records = []
                for _, row in processed_df.iterrows():
                    row_data = {key: self.convert_to_json_serializable(value) 
                               for key, value in row.to_dict().items()}
                    
                    def get_val(col_name, default=None):
                        return row.get(col_name, default) if col_name in row and pd.notna(row[col_name]) else default
                    
                    def get_date(col_name):
                        val = get_val(col_name)
                        if val is None:
                            return None
                        if isinstance(val, (pd.Timestamp, datetime)):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None
                    
                    # Get values
                    employee_id_val = get_val('employee_id')
                    # Convert employee_id to string if it exists, or None
                    if employee_id_val is not None:
                        employee_id_str = str(employee_id_val)
                    else:
                        employee_id_str = None
                    
                    name = str(get_val('name', ''))
                    position = str(get_val('position', ''))
                    role = str(get_val('role', position))
                    hourly_rate = float(get_val('hourly_rate', 0.0))
                    hours_worked = float(get_val('hours_worked', 0.0))
                    total_wages = float(get_val('total_wages', 0.0))
                    overtime_hours = float(get_val('overtime_hours', 0.0))
                    
                    # Debug logging
                    if len(records) < 3:
                        logging.info(f"Row {len(records)}: Employee={name}, ID={employee_id_str}, Hours={hours_worked}, Wages=${total_wages}")
                    
                    records.append((
                        restaurant_id,                           # 1
                        filename,                                # 2
                        json.dumps(row_data),                   # 3 - data
                        get_date('date'),                       # 4 - date
                        employee_id_str,                        # 5 - employee_id (as string)
                        name,                                    # 6 - employee_name
                        name,                                    # 7 - name
                        position,                                # 8 - position
                        role,                                    # 9 - role
                        hourly_rate,                            # 10 - hourly_rate
                        hours_worked,                           # 11 - hours_worked
                        total_wages,                            # 12 - total_wages
                        overtime_hours,                         # 13 - overtime_hours
                        get_date('hire_date'),                  # 14 - hire_date
                        get_date('termination_date')            # 15 - termination_date
                    ))
                
                cur.executemany(query, records)
                self.conn.commit()
                logging.info(f"✅ Successfully inserted {len(records)} employee records into the database.")
                logging.info(f"📊 Total hours: {sum(r[11] for r in records):.2f}")
                logging.info(f"📊 Total wages: ${sum(r[12] for r in records):.2f}")
                
        except Exception as e:
            self.conn.rollback()
            logging.error(f"❌ Error inserting employee data: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
# Main upload function
async def upload_file(
        file: UploadFile,
        restaurant_name: str,
        category: str,
        current_user: dict
) -> Dict[str, Any]:
    """Upload a file (PDF, XLS, DOCX, or CSV) to S3"""
    try:
        restaurant = await verify_restaurant_access(restaurant_name, current_user)

        file_type = await validate_file(file)
        # ============ NEW: CSV COLUMN VALIDATION ============
        if file_type == "csv":
            await file.seek(0)
            file_content_for_validation = await file.read()
            
            # Validate CSV columns
            is_valid, validation_result = validate_csv_columns(
                file_content=file_content_for_validation,
                category=category,
                filename=file.filename
            )
            
            if not is_valid:
                logger.warning(f"CSV validation failed for {file.filename}: {validation_result.get('message')}")
                
                # Return detailed error with requirements
                error_response = {
                    "error": "CSV Validation Failed",
                    "message": validation_result.get('message'),
                    "filename": file.filename,
                    "category": category,
                    "found_columns": validation_result.get('found_columns', []),
                    "missing_required": validation_result.get('missing_required', []),
                    "missing_at_least_one": validation_result.get('missing_at_least_one', []),
                    "example": validation_result.get('example_fix', {}),
                    "requirements": get_column_requirements(category)
                }
                
                raise HTTPException(
                    status_code=400,
                    detail=error_response
                )
            
            logger.info(f"✅ CSV validation passed for {file.filename}")
            logger.info(f"Found columns: {validation_result.get('found_columns')}")
            
            # Reset file pointer after validation
            await file.seek(0)


         # ============ END: CSV COLUMN VALIDATION ============

        s3_folder = get_user_restaurant_path(restaurant_name, current_user["id"], category)

        #random_id = uuid.uuid4().hex[:8]
        #filename_base = os.path.splitext(file.filename)[0]
        #new_filename = f"{filename_base}_{random_id}.{file.filename.split('.')[-1]}"
        new_filename = file.filename
        s3_key = f"{s3_folder}/{new_filename}"

        if file_type == "csv":
            row_count, column_count = await process_csv(file)
        elif file_type in ["xls", "xlsx"]:
            row_count, column_count = await process_xls(file)
        elif file_type == "docx":
            row_count, column_count = await process_docx(file)
        elif file_type == "pdf":
            row_count, column_count = await process_pdf(file)

        await file.seek(0)
        file_content = await file.read()

        content_type = get_content_type(file.filename)

        if not file_content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")

        filename = new_filename
        restaurant_id = get_restaurant_id_by_name(restaurant_name)
        
        # Process CSV files
        if file_type == "csv":
            class MockUploadFile:
                def __init__(self, content, fname):
                    self.content = content
                    self.filename = fname
                async def seek(self, pos): 
                    self.content.seek(pos)
                async def read(self):
                    return self.content.read()
            
            csv_conn = get_db_connection()
            try:
                logging.info(f"Processing CSV file for category: {category}, restaurant_id: {restaurant_id}")
                
                file_for_processing = io.BytesIO(file_content)
                mock_file = MockUploadFile(file_for_processing, file.filename)
                
                csvs_processor = SimplifiedCSVProcessor()
                csvs_uploader = CSVUploader(csv_conn)
                
                if category == 'Sales':
                    data = await csvs_processor.process_sales_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_sales(data, restaurant_id, filename)
                        logging.info(f"✓ Sales data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid sales data to insert")

                elif category == 'Labor':
                    data = await csvs_processor.process_employees_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_employees(data, restaurant_id, filename)
                        logging.info(f"✓ Employee data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid employee data to insert")

                elif category == 'Inventory':
                    data = await csvs_processor.process_inventory_csv(mock_file, restaurant_id, csv_conn)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_inventory(data, restaurant_id, filename)
                        logging.info(f"✅ Inventory data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"⚠️ No valid inventory data to insert")

                elif category == 'Menu':
                    data = await csvs_processor.process_menu_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_menu(data, restaurant_id, filename)
                        logging.info(f"✓ Menu data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid menu data to insert")

                else:
                    logging.warning(f"Unknown category provided: {category}")

            except Exception as e:
                logging.error(f"Failed to handle CSV upload for category '{category}': {str(e)}")
                logging.error(f"Traceback: {traceback.format_exc()}")
            finally:
                return_db_connection(csv_conn)

        # Upload to S3
        try:
            s3_client.put_object(
                Bucket=BUCKET_NAME,
                Key=s3_key,
                Body=file_content,
                ContentType=content_type,
            )

            logger.info(f"✓ File uploaded successfully to S3: {s3_key}")

            return {
                "message": "File uploaded successfully",
                "filename": new_filename,
                "original_filename": file.filename,
                "restaurant": restaurant_name,
                "category": category,
                "s3_key": s3_key,
                "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{s3_key}",
                "rows": row_count,
                "columns": column_count
            }

        except Exception as e:
            logger.error(f"Error uploading to S3: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error uploading file to storage: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in upload_file: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error processing upload: {str(e)}")


async def upload_file_id(
        file: UploadFile,
        restaurant_id: int,
        category: str,
        current_user: dict
) -> Dict[str, Any]:
    """Upload a file (PDF, XLS, DOCX, or CSV) to S3 using restaurant ID"""
    try:
        restaurant = await verify_restaurant_access_id(restaurant_id, current_user)

        file_type = await validate_file(file)
        # ============= NEW: CSV COLUMN VALIDATION =============
        if file_type == "csv":
            await file.seek(0)
            file_content_for_validation = await file.read()
            
            # Validate CSV columns
            is_valid, validation_result = validate_csv_columns(
                file_content=file_content_for_validation,
                category=category,
                filename=file.filename
            )
            
            if not is_valid:
                logger.warning(f"CSV validation failed for {file.filename}: {validation_result.get('message')}")
                
                # Return detailed error with requirements
                error_response = {
                    "error": "CSV Validation Failed",
                    "message": validation_result.get('message'),
                    "filename": file.filename,
                    "category": category,
                    "found_columns": validation_result.get('found_columns', []),
                    "missing_required": validation_result.get('missing_required', []),
                    "missing_at_least_one": validation_result.get('missing_at_least_one', []),
                    "example": validation_result.get('example_fix', {}),
                    "requirements": get_column_requirements(category)
                }
                
                raise HTTPException(
                    status_code=400,
                    detail=error_response
                )
            
            logger.info(f"✅ CSV validation passed for {file.filename}")
            logger.info(f"Found columns: {validation_result.get('found_columns')}")
            
            # Reset file pointer after validation
            await file.seek(0)


        # ============= END: CSV COLUMN VALIDATION =============

        s3_folder = get_user_restaurant_path_id(restaurant_id, current_user["id"], category)

        random_id = uuid.uuid4().hex[:8]
        filename_base = os.path.splitext(file.filename)[0]
        new_filename = f"{filename_base}_{random_id}.{file.filename.split('.')[-1]}"
        s3_key = f"{s3_folder}/{new_filename}"

        if file_type == "csv":
            row_count, column_count = await process_csv(file)
        elif file_type in ["xls", "xlsx"]:
            row_count, column_count = await process_xls(file)
        elif file_type == "docx":
            row_count, column_count = await process_docx(file)
        elif file_type == "pdf":
            row_count, column_count = await process_pdf(file)

        await file.seek(0)
        file_content = await file.read()

        content_type = get_content_type(file.filename)

        if not file_content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")

        filename = new_filename
        restaurant_name = get_restaurant_name_by_id(restaurant_id)

        # Process CSV files
        if file_type == "csv":
            class MockUploadFile:
                def __init__(self, content, fname):
                    self.content = content
                    self.filename = fname
                async def seek(self, pos): 
                    self.content.seek(pos)
                async def read(self):
                    return self.content.read()
            
            csv_conn = get_db_connection()
            try:
                logging.info(f"Processing CSV file for category: {category}, restaurant_id: {restaurant_id}")
                
                file_for_processing = io.BytesIO(file_content)
                mock_file = MockUploadFile(file_for_processing, file.filename)
                
                csvs_processor = SimplifiedCSVProcessor()
                csvs_uploader = CSVUploader(csv_conn)
                
                if category == 'Sales':
                    data = await csvs_processor.process_sales_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_sales(data, restaurant_id, filename)
                        logging.info(f"✓ Sales data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid sales data to insert")

                elif category == 'Labor':
                    data = await csvs_processor.process_employees_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_employees(data, restaurant_id, filename)
                        logging.info(f"✓ Employee data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid employee data to insert")

                elif category == 'Inventory':
                    data = await csvs_processor.process_inventory_csv(mock_file, restaurant_id, csv_conn)
                    #data = await csvs_processor.process_inventory_csv(mock_file)
                    #data = await csvs_processor.process_inventory_csv(mock_file, restaurant_id, csv_conn)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_inventory(data, restaurant_id, filename)
                        logging.info(f"✅ Inventory data inserted: {len(data)} rows")
                    #if data is not None and not data.empty:
                    #    csvs_uploader.insert_inventory(data, restaurant_id, filename)
                    #    logging.info(f"✅ Inventory data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"⚠️ No valid inventory data to insert")

                elif category == 'Menu':
                    data = await csvs_processor.process_menu_csv(mock_file)
                    if data is not None and not data.empty:
                        csvs_uploader.insert_menu(data, restaurant_id, filename)
                        logging.info(f"✓ Menu data inserted: {len(data)} rows")
                    else:
                        logging.warning(f"✗ No valid menu data to insert")

                else:
                    logging.warning(f"Unknown category provided: {category}")

            except Exception as e:
                logging.error(f"Failed to handle CSV upload for category '{category}': {str(e)}")
                logging.error(f"Traceback: {traceback.format_exc()}")
            finally:
                return_db_connection(csv_conn)

        # Upload to S3
        try:
            s3_client.put_object(
                Bucket=BUCKET_NAME,
                Key=s3_key,
                Body=file_content,
                ContentType=content_type,
            )

            logger.info(f"✓ File uploaded successfully to S3: {s3_key}")

            return {
                "message": "File uploaded successfully",
                "filename": new_filename,
                "original_filename": file.filename,
                "restaurant": restaurant_name,
                "category": category,
                "s3_key": s3_key,
                "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{s3_key}",
                "rows": row_count,
                "columns": column_count
            }

        except Exception as e:
            logger.error(f"Error uploading to S3: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error uploading file to storage: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in upload_file_id: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error processing upload: {str(e)}")
    


async def list_cat_csv_files(
        restaurant_name: str,
        category: str,
        current_user: dict
) -> List[Dict[str, Any]]:
    """List all files for a restaurant category"""
    try:
        # Verify restaurant access
        restaurant = await verify_restaurant_access(restaurant_name, current_user)

        # Generate S3 path
        s3_folder = get_user_restaurant_path(restaurant_name, current_user["id"], category)

        # List objects in the folder
        try:
            response = s3_client.list_objects_v2(
                Bucket=BUCKET_NAME,
                Prefix=f"{s3_folder}/"
            )

            files = []

            if 'Contents' in response:
                for obj in response['Contents']:
                    # Skip folders themselves
                    if obj['Key'].endswith('/'):
                        continue

                    filename = os.path.basename(obj['Key'])

                    # Only include files
                    if filename:
                        files.append({
                            "filename": filename,
                            "size": obj['Size'],
                            "last_modified": obj['LastModified'].strftime("%Y-%m-%d %H:%M:%S"),
                            "s3_key": obj['Key'],
                            "category": str(obj['Key']).split("/")[-2],
                            "url": f"https://{BUCKET_NAME}.s3.amazonaws.com/{obj['Key']}"
                        })

            return {
                "restaurant": restaurant_name,
                "category": category,
                "files": files,
                "count": len(files)
            }

        except Exception as e:
            logger.error(f"Error listing files from S3: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Error retrieving file list: {str(e)}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in list_cat_csv_files: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error listing CSV files: {str(e)}"
        )